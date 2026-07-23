import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("F:/Vanya/fort_calculator.xlsx", data_only=False)

output_path = "F:/Vanya/blindage_formulas.txt"
with open(output_path, "w", encoding="utf-8") as f:
    # Вытащим формулы листа Блиндажи_и_Укрытия
    f.write("=== EXCEL FORMULAS FOR BLINDAGES ===\n")
    sheet = wb['Блиндажи_и_Укрытия']
    for r in range(4, 51):
        vals = []
        for c in range(1, 8):
            cell = sheet.cell(row=r, column=c)
            val = str(cell.value) if cell.value is not None else ""
            vals.append(f"{cell.coordinate}: {val}")
        f.write(" | ".join(vals) + "\n")
        
    # Вытащим также формулы листа Укрытия_техники
    f.write("\n=== EXCEL FORMULAS FOR VEHICLE SHELTERS ===\n")
    sheet_tech = wb['Укрытия_техники']
    for r in range(4, 48):
        vals = []
        for c in range(1, 8):
            cell = sheet_tech.cell(row=r, column=c)
            val = str(cell.value) if cell.value is not None else ""
            vals.append(f"{cell.coordinate}: {val}")
        f.write(" | ".join(vals) + "\n")

print(f"Blindage and Tech formulas written to {output_path}")
