import openpyxl
import sys

# Настроим кодировку stdout для безопасности
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("F:/Vanya/fort_calculator.xlsx", data_only=True)
output_path = "F:/Vanya/inspect_excel.txt"

with open(output_path, "w", encoding="utf-8") as f:
    f.write("=== INSPECT EXCEL ===\n")
    f.write(f"Sheets in workbook: {wb.sheetnames}\n\n")
    
    for name in wb.sheetnames:
        sheet = wb[name]
        f.write(f"=========================================\n")
        f.write(f"Sheet: {name}\n")
        f.write(f"Dimensions: max_row={sheet.max_row}, max_column={sheet.max_column}\n")
        f.write(f"=========================================\n")
        
        row_idx = 0
        for row in sheet.iter_rows(values_only=False):
            row_idx += 1
            # Считываем ячейки (значения и формулы)
            vals = []
            for cell in row:
                if cell.value is not None:
                    # Запишем значение и формулу, если есть
                    val_str = str(cell.value)
                    vals.append(f"Col {cell.column} ({cell.coordinate}): {val_str}")
            if vals:
                f.write(f"Row {row_idx}: " + " | ".join(vals) + "\n")
        f.write("\n\n")

print(f"Excel inspected. Output written to {output_path}")
